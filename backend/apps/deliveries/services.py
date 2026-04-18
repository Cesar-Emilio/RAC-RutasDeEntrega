"""
services.py — Lógica de negocio para el procesamiento de rutas

Flujo completo:
    process_route(route_id)
        └── 1. parse_input_file()     - extrae filas del CSV/JSON
        └── 2. resolve_coordinates()  - geocodifica las filas que no traen lat/lng
        └── 3. save_delivery_points() - persiste los DeliveryPoint en BD
        └── 4. optimize_route()       - calcula el orden óptimo
        └── 5. save_solution()        - persiste RouteSolution + RouteSolutionDetail
"""

import csv
import json
import io
import math
import time
import requests
import openpyxl
from django.conf import settings
from typing import TypedDict

from django.db import transaction

from config.logging_utils import get_logger

logger = get_logger(__name__)

from .models import (
    DeliveryPoint,
    FileType,
    Route,
    RouteSolution,
    RouteSolutionDetail,
    Status,
)

class RawRow(TypedDict):
    """Fila normalizada que sale del parser, antes de geocodificar."""
    address: str
    latitude: float | None
    longitude: float | None
    receiver_name: str
    package_quantity: int

class ResolvedRow(TypedDict):
    """Fila con coordenadas garantizadas, lista para persistir."""
    address: str
    latitude: float
    longitude: float
    receiver_name: str
    package_quantity: int

class RouteProcessingError(Exception):
    """
    Error controlado durante el procesamiento de una ruta.
    El task lo captura para marcar la ruta con Status.ERROR
    sin propagar un crash genérico.
    """

def parse_input_file(route: Route) -> list[RawRow]:
    """
    Lee el archivo asociado a la ruta y devuelve una lista de RawRow.

    Formatos soportados
    -------------------
    CSV — debe tener columna 'address'. Las columnas 'latitude' y 'longitude'
          son opcionales; si no están presentes se dejan en None.

    JSON — acepta dos estructuras:
        · Lista de objetos: [{"address": "...", "latitude": ..., "longitude": ...}, ...]
        · Objeto con clave "deliveries": {"deliveries": [...]}

    Raises
    ------
    RouteProcessingError
        Si el archivo no tiene el formato esperado o faltan columnas requeridas.
    """
    logger.debug(
        "parse_input_file | action=start | route_id={route_id}",
        route_id=route.id,
    )
    try:
        input_file = route.input_file
    except Route.input_file.RelatedObjectDoesNotExist:
        raise RouteProcessingError(f"La ruta {route.id} no tiene archivo de entrada asociado.")

    file_obj = input_file.file
    file_obj.seek(0)
    content = file_obj.read()

    # Obtiene solo el archivo y lo procesa según su tipo
    if input_file.file_type == FileType.CSV:
        rows = parse_csv(content)
    elif input_file.file_type == FileType.JSON:
        rows = parse_json(content)
    elif input_file.file_type == FileType.XLSX:
        rows = parse_xlsx(content)
    else:
        raise RouteProcessingError(f"Tipo de archivo no soportado: {input_file.file_type}")

    logger.debug(
        "parse_input_file | action=end | route_id={route_id} | file_type={file_type} | rows_parsed={rows_parsed}",
        route_id=route.id,
        file_type=input_file.file_type,
        rows_parsed=len(rows),
    )
    return rows

def parse_csv(content: bytes) -> list[RawRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    required = {"address", "receiver_name", "package_quantity"}
    missing = required - set(reader.fieldnames or [])
    if missing:
        raise RouteProcessingError(f"El CSV debe contener las columnas: {', '.join(missing)}.")

    rows: list[RawRow] = []
    for i, raw in enumerate(reader, start=2):
        address = raw.get("address", "").strip()
        if not address:
            raise RouteProcessingError(f"Fila {i}: la columna 'address' está vacía.")
        
        receiver_name = raw.get("receiver_name", "").strip()
        if not receiver_name:
            raise RouteProcessingError(f"Fila {i}: la columna 'receiver_name' está vacía.")

        package_quantity = _to_int(raw.get("package_quantity"))
        if package_quantity is None or package_quantity < 1:
            raise RouteProcessingError(f"Fila {i}: 'package_quantity' debe ser un entero positivo.")

        rows.append({
            "address": address,
            "latitude": _to_float(raw.get("latitude")),
            "longitude": _to_float(raw.get("longitude")),
            "receiver_name": receiver_name,
            "package_quantity": package_quantity,
        })

    if not rows:
        raise RouteProcessingError("El CSV no contiene filas de datos.")
    return rows

def parse_json(content: bytes) -> list[RawRow]:
    try:
        data = json.loads(content.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise RouteProcessingError(f"El JSON no es válido: {exc}")

    # Acepta lista directa o dict con clave "deliveries"
    if isinstance(data, dict):
        data = data.get("deliveries", [])
    if not isinstance(data, list) or not data:
        raise RouteProcessingError("El JSON debe ser una lista de objetos o contener la clave 'deliveries'.")

    rows: list[RawRow] = []
    for i, item in enumerate(data):
        address = str(item.get("address", "")).strip()
        if not address:
            raise RouteProcessingError(f"Elemento {i}: falta el campo 'address'.")
        
        receiver_name = str(item.get("receiver_name", "")).strip()
        if not receiver_name:
            raise RouteProcessingError(f"Elemento {i}: falta el campo 'receiver_name'.")

        package_quantity = _to_int(item.get("package_quantity"))
        if package_quantity is None or package_quantity < 1:
            raise RouteProcessingError(f"Elemento {i}: 'package_quantity' debe ser un entero positivo.")

        rows.append({
            "address": address,
            "latitude": _to_float(item.get("latitude")),
            "longitude": _to_float(item.get("longitude")),
            "receiver_name": receiver_name,
            "package_quantity": package_quantity,
        })
    return rows

def parse_xlsx(content: bytes) -> list[RawRow]:
    """
    Parsea un archivo Excel siguiendo la plantilla oficial.

    Estructura esperada
    -------------------
    Hoja: "Entregas" (se ignora la de instrucciones)
    Columna: address
    Columna: latitude
    Columna: longitude

    Raises
    ------
    RouteProcessingError
        Si falta la hoja "Entregas", la columna address, o alguna fila la tiene vacía.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise RouteProcessingError(f"El archivo Excel no es válido: {exc}")

    if "Entregas" not in wb.sheetnames:
        raise RouteProcessingError(
            "El archivo Excel debe contener una hoja llamada 'Entregas'."
        )

    ws = wb["Entregas"]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        raise RouteProcessingError("La hoja 'Entregas' está vacía.")

    required = {"address", "receiver_name", "package_quantity"}
    missing = required - set(headers)
    if missing:
        raise RouteProcessingError(
            f"La hoja 'Entregas' debe contener las columnas: {', '.join(missing)}."
        )

    idx_address        = headers.index("address")
    idx_latitude       = headers.index("latitude") if "latitude" in headers else None
    idx_longitude      = headers.index("longitude") if "longitude" in headers else None
    idx_receiver       = headers.index("receiver_name")
    idx_pkg_quantity   = headers.index("package_quantity")

    rows: list[RawRow] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if all(cell is None for cell in row):
            continue
    
        address = str(row[idx_address]).strip() if row[idx_address] is not None else ""
        if not address:
            raise RouteProcessingError(f"Fila {row_num}: la columna 'address' está vacía.")

        receiver_name = str(row[idx_receiver]).strip() if row[idx_receiver] is not None else ""
        if not receiver_name:
            raise RouteProcessingError(f"Fila {row_num}: la columna 'receiver_name' está vacía.")

        package_quantity = _to_int(row[idx_pkg_quantity])
        if package_quantity is None or package_quantity < 1:
            raise RouteProcessingError(f"Fila {row_num}: 'package_quantity' debe ser un entero positivo.")

        rows.append({
            "address": address,
            "latitude":  _to_float(row[idx_latitude])  if idx_latitude  is not None else None,
            "longitude": _to_float(row[idx_longitude]) if idx_longitude is not None else None,
            "receiver_name": receiver_name,
            "package_quantity": package_quantity,
        })

    if not rows:
        raise RouteProcessingError("La hoja 'Entregas' no contiene filas de datos.")

    return rows

def _to_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def _to_int(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

def resolve_coordinates(rows: list[RawRow], route_id: int | None = None) -> list[ResolvedRow]:
    """
    Garantiza que todas las filas tengan lat/lng.

    - Si la fila ya tiene coordenadas -> se usa directamente.
    - Si le faltan -> se llama a geocode_address().

    Raises
    ------
    RouteProcessingError
        Si la geocodificación falla para alguna dirección.
    """
    resolved: list[ResolvedRow] = []
    geocoded_count = 0

    for row in rows:
        if row["latitude"] is not None and row["longitude"] is not None:
            lat, lng = row["latitude"], row["longitude"]
        else:
            logger.debug(
                "resolve_coordinates | action=geocode_start | route_id={route_id} | address={address}",
                route_id=route_id,
                address=row["address"],
            )
            lat, lng = geocode_address(row["address"])
            geocoded_count += 1
            logger.debug(
                "resolve_coordinates | action=geocode_end | route_id={route_id} "
                "| address={address} | lat={lat} | lng={lng}",
                route_id=route_id,
                address=row["address"],
                lat=lat,
                lng=lng,
            )

        resolved.append({
            "address": row["address"],
            "latitude": lat,
            "longitude": lng,
            "receiver_name": row["receiver_name"],
            "package_quantity": row["package_quantity"],
        })

    if geocoded_count:
        logger.debug(
            "resolve_coordinates | action=summary | route_id={route_id} "
            "| total_rows={total} | geocoded={geocoded} | pre_resolved={pre_resolved}",
            route_id=route_id,
            total=len(rows),
            geocoded=geocoded_count,
            pre_resolved=len(rows) - geocoded_count,
        )
    return resolved

def geocode_address(address: str) -> tuple[float, float]:
    """
    Convierte una dirección de texto en coordenadas (lat, lng).

    Raises
    ------
    RouteProcessingError
        Si la API no devuelve resultados o falla la conexión.
    """
    url = "https://nominatim.openstreetmap.org/search"

    params = {
        "q": address,
        "format": "json",
        "limit": 1
    }
    
    headers = {
        "User-Agent": "route-optimizer-app"
    }

    start = time.perf_counter()
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=5)
        resp.raise_for_status()
        results = resp.json()
    except requests.Timeout:
        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.warning(
            "geocode_address | action=timeout | address={address} | elapsed_ms={elapsed_ms}",
            address=address,
            elapsed_ms=elapsed_ms,
        )
        raise RouteProcessingError(f"Timeout al conectar con Nominatim para: '{address}'")
    except requests.RequestException as exc:
        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.error(
            "geocode_address | action=request_error | address={address} "
            "| error={error} | elapsed_ms={elapsed_ms}",
            address=address,
            error=str(exc),
            elapsed_ms=elapsed_ms,
        )
        raise RouteProcessingError(f"Error al conectar con Nominatim: {exc}")

    if not results:
        logger.warning(
            "geocode_address | action=no_results | address={address}",
            address=address,
        )
        raise RouteProcessingError(f"No se encontraron coordenadas para: '{address}'")

    return float(results[0]["lat"]), float(results[0]["lon"])

def save_delivery_points(route: Route, rows: list[ResolvedRow]) -> list[DeliveryPoint]:
    """
    Crea los DeliveryPoint en BD usando bulk_create para minimizar queries.
    Devuelve la lista de instancias creadas en el mismo orden que 'rows'.
    """
    points = [
        DeliveryPoint(
            route=route,
            address=row["address"],
            latitude=row["latitude"],
            longitude=row["longitude"],
            receiver_name=row["receiver_name"],
            package_quantity=row["package_quantity"],
        )
        for row in rows
    ]
    DeliveryPoint.objects.bulk_create(points)
    saved = list(DeliveryPoint.objects.filter(route=route).order_by("id"))
    logger.debug(
        "save_delivery_points | action=bulk_create | route_id={route_id} | points_saved={count}",
        route_id=route.id,
        count=len(saved),
    )
    return saved

def optimize_route(
    points: list[DeliveryPoint],
    warehouse_lat: float,
    warehouse_lng: float,
    k_opt: int = 0,
) -> tuple[list[DeliveryPoint], float]:
    """
    Calcula el orden óptimo de visita para los puntos de entrega.
 
    Algoritmo
    ---------
    1. Christofides simplificado (MST + matching de nodos de grado impar
       por nearest-neighbor) como heurística de construcción inicial.
       Garantiza una solución ≤ 1.5× el óptimo en grafos métricos.
    2. k-opt improvement: ejecuta _two_opt exactamente k_opt veces, 
       se eliminó el threshold. Cada iteración parte del tour
       mejorado de la anterior, acumulando mejoras sucesivas.
       k_opt=0 omite la mejora completamente.
 
    Parámetros
    ----------
    points         : DeliveryPoints a ordenar (sin el warehouse).
    warehouse_lat  : Latitud del almacén (nodo origen y destino).
    warehouse_lng  : Longitud del almacén.
    k_opt          : Número de pasadas de mejora 2-opt (0 = sin mejora).
 
    Retorna
    -------
    ordered_points : list[DeliveryPoint] en el orden óptimo calculado.
    total_distance : Distancia total en km incluyendo salida y regreso al warehouse.
    """
    n_points = len(points)
    start = time.perf_counter()

    logger.debug(
        "optimize_route | action=start | points={n_points} | warehouse_lat={lat} | warehouse_lng={lng}",
        n_points=n_points,
        lat=warehouse_lat,
        lng=warehouse_lng,
    )

    # Caso 1: no hay paquetes a entregar
    if not points:
        raise RouteProcessingError("No hay puntos de entrega para optimizar.")
    # Caso 2: Solo hay un paquete a entregar:
    # 1. Calcula la distancia del almacen a ese punto
    # 2. Duplica esa distancia (ida y vuelta)
    # 3. Devuelve la misma lista inicial, no se ocupa ordenamiento
    if len(points) == 1:
        d = haversine(warehouse_lat, warehouse_lng,
                       float(points[0].latitude), float(points[0].longitude))
        total_distance = round(d * 2, 4)
        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.debug(
            "optimize_route | action=end | strategy=single_point | total_distance_km={dist} "
            "| execution_time_ms={elapsed_ms}",
            dist=total_distance,
            elapsed_ms=elapsed_ms,
        )
        return points, total_distance
 
    # La matriz incluye el warehouse en el índice 0, los puntos en 1..n
    all_coords = [(warehouse_lat, warehouse_lng)] + [
        (float(p.latitude), float(p.longitude)) for p in points
    ]
    n_total = len(all_coords)  # Cantidad de paquetes a entregar + 1 (el almacén de partida)
 
    # Se crea una matriz [i][j] con la distancia entre cualquier par de nodos
    # Es una matriz simétrica
    graph = build_distance_matrix(all_coords)
 
    # Se construye un tour inicial siendo un recorrido que inicia y acaba en el almacén
    logger.debug(
        "optimize_route | action=christofides_start | nodes={n_total}",
        n_total=n_total,
    )
    tour = christofides_tour(graph, n_total)
    logger.debug(
        "optimize_route | action=christofides_end | tour_length={tour_length}",
        tour_length=len(tour),
    )
 
    # Si hay 50 paquetes o menos, aplica 2-opt
    if k_opt > 0:
        logger.debug(
            "optimize_route | action=two_opt_start | points={n_points} | k_opt={k_opt}",
            n_points=n_points,
            k_opt=k_opt
        )
        for iteration in range(k_opt):
            tour_before = _tour_distance(tour, graph)
            tour = two_opt(tour, graph)
            tour_after = _tour_distance(tour, graph)
            logger.debug(
                "optimize_route | action=two_opt_improvement | iteration={iteration}/{k_opt} | tour_before={tour_before} | tour_after={tour_after} | improvement = {improvement}",
                iteration=iteration + 1,
                k_opt=k_opt,
                tour_before=tour_before,
                tour_after=tour_after,
                improvement=tour_after - tour_before,
            )
            # Early stopping: si la iteración no mejoró nada, las siguientes tampoco lo harán
            if tour_before - tour_after < 1e-10:
                logger.debug(
                    "optimize_route | action=two_opt_stop | iteration={iteration}/{k_opt}",
                    iteration=iteration + 1,
                    k_opt=k_opt,
                )
                break
        logger.debug("optimize_route | action=two_opt_end")
    else:
        logger.debug(
            "optimize_route | action=two_opt_skipped | points={n_points} | k_opt={k_opt}",
            n_points=n_points,
            k_opt=k_opt
        )
 
    # Suma todas las aristas del tour final
    total_distance = _tour_distance(tour, graph)
 
    # tour[0] y tour[-1] son el almacén; los puntos están en 1..n
    ordered_indices = [i - 1 for i in tour[1:-1]]  # descarta el almacén, convierte a índice de points[]
    ordered_points = [points[i] for i in ordered_indices]

    elapsed_ms = int((time.perf_counter() - start) * 1000)
    logger.debug(
        "optimize_route | action=end | strategy=christofides | total_distance_km={dist} "
        "| execution_time_ms={elapsed_ms}",
        dist=round(total_distance, 4),
        elapsed_ms=elapsed_ms,
    )
 
    return ordered_points, round(total_distance, 4)


def build_distance_matrix(coords: list[tuple[float, float]]) -> list[list[float]]:
    """
    Matriz simétrica de distancias Haversine entre todos los nodos.
    coords[0] = warehouse, coords[1..n] = delivery points.
 
    Punto de extensión: reemplazar haversine() por una API de distancias
    viales (OSRM, Google Distance Matrix) para rutas más realistas.
    """
    n = len(coords)
    matrix = [[0.0] * n for _ in range(n)]
    
    for i in range(n):
        for j in range(i + 1, n):
            d = haversine(coords[i][0], coords[i][1], coords[j][0], coords[j][1])
            matrix[i][j] = d
            matrix[j][i] = d
    
    return matrix

def christofides_tour(graph: list[list[float]], n: int) -> list[int]:
    """
    Heurística de Christofides simplificada para TSP métrico.
 
    Pasos
    -----
    1. MST de Prim sobre el grafo completo (nodos 0..n-1).
    2. Identificar nodos de grado impar en el MST.
    3. Matching mínimo aproximado de los nodos impares
       (greedy nearest-neighbor; el matching exacto requiere Blossom V).
    4. Construir multigrafo MST + matching y extraer circuito euleriano.
    5. Shortcutting: saltar nodos ya visitados para obtener tour hamiltoniano.
 
    El tour comienza y termina en el nodo 0 (warehouse).
    """
    mst_adj = prim_mst(graph, n)
    odd_nodes = odd_degree_nodes(mst_adj, n)
    matching_edges = greedy_matching(odd_nodes, graph)
 
    multigraph: list[list[int]] = [[] for _ in range(n)]
    for u, neighbors in enumerate(mst_adj):
        for v in neighbors:
            multigraph[u].append(v)
    for u, v in matching_edges:
        multigraph[u].append(v)
        multigraph[v].append(u)
 
    euler = eulerian_circuit(multigraph, start=0)
    tour = shortcut(euler)
    return tour
 
 
def prim_mst(graph: list[list[float]], n: int) -> list[list[int]]:
    """
    Genera el MST con algoritmo de Prim.
    Retorna lista de adyacencia (sin pesos).
    """
    in_mst = [False] * n
    min_edge = [math.inf] * n
    parent = [-1] * n
    min_edge[0] = 0.0
 
    adj: list[list[int]] = [[] for _ in range(n)]
 
    for _ in range(n):
        u = min((v for v in range(n) if not in_mst[v]), key=lambda v: min_edge[v])
        in_mst[u] = True
 
        if parent[u] != -1:
            adj[parent[u]].append(u)
            adj[u].append(parent[u])
 
        for v in range(n):
            if not in_mst[v] and graph[u][v] < min_edge[v]:
                min_edge[v] = graph[u][v]
                parent[v] = u
 
    return adj

def odd_degree_nodes(adj: list[list[int]], n: int) -> list[int]:
    """Retorna los índices de nodos con grado impar en el árbol dado."""
    return [i for i in range(n) if len(adj[i]) % 2 != 0]
 
def greedy_matching(nodes: list[int], graph: list[list[float]]) -> list[tuple[int, int]]:
    """
    Matching mínimo aproximado sobre un conjunto de nodos.
    Greedy: empareja cada nodo libre con su vecino libre más cercano.
    """
    remaining = list(nodes)
    edges: list[tuple[int, int]] = []
 
    while len(remaining) >= 2:
        u = remaining[0]
        best_v = min(remaining[1:], key=lambda v: graph[u][v])
        edges.append((u, best_v))
        remaining.remove(u)
        remaining.remove(best_v)
 
    return edges

def eulerian_circuit(adj: list[list[int]], start: int) -> list[int]:
    """
    Circuito euleriano por el algoritmo de Hierholzer (iterativo).
    Modifica una copia de adj para no alterar el multigrafo original.
    """
    graph_copy = [list(neighbors) for neighbors in adj]
    stack = [start]
    circuit: list[int] = []

    while stack:
        v = stack[-1]
        if graph_copy[v]:
            u = graph_copy[v].pop()
            graph_copy[u].remove(v)
            stack.append(u)
        else:
            circuit.append(stack.pop())

    circuit.reverse()
    return circuit
 
 
def shortcut(euler: list[int]) -> list[int]:
    """
    Convierte el circuito euleriano en tour hamiltoniano
    saltando nodos ya visitados (shortcutting).
    Preserva el nodo inicial al inicio y al final.
    """
    seen: set[int] = set()
    tour: list[int] = []
    for node in euler:
        if node not in seen:
            seen.add(node)
            tour.append(node)
    tour.append(tour[0])  # cerrar el ciclo
    return tour
 
def two_opt(tour: list[int], graph: list[list[float]]) -> list[int]:
    """
    Mejora el tour intercambiando pares de aristas (2-opt).
 
    Itera hasta que ningún intercambio reduzca la distancia total.
    Complejidad O(n²) por pasada; en la práctica converge en pocas iteraciones.
 
    El almacen (tour[0] = tour[-1]) se mantiene fijo como punto de inicio/fin.
    Solo se permutan los nodos internos (índices 1..n-2 del tour).
    """
    best = tour[:]
    improved = True
 
    while improved:
        improved = False
        for i in range(1, len(best) - 2):
            for j in range(i + 1, len(best) - 1):
                current = graph[best[i - 1]][best[i]] + graph[best[j]][best[j + 1]]
                swapped = graph[best[i - 1]][best[j]] + graph[best[i]][best[j + 1]]
 
                if swapped < current - 1e-10:
                    best[i:j + 1] = best[i:j + 1][::-1]
                    improved = True
 
    return best

def _tour_distance(tour: list[int], graph: list[list[float]]) -> float:
    """Suma la distancia total de un tour cerrado."""
    return sum(graph[tour[i]][tour[i + 1]] for i in range(len(tour) - 1))

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distancia entre dos puntos geográficos en kilómetros (fórmula Haversine)."""
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))

def save_solution(route: Route, ordered_points: list[DeliveryPoint], total_distance: float) -> RouteSolution:
    """
    Persiste RouteSolution y sus RouteSolutionDetail en una transacción atómica.
    """
    with transaction.atomic():
        solution = RouteSolution.objects.create(
            route=route,
            total_distance=total_distance,
        )
        details = [
            RouteSolutionDetail(
                solution=solution,
                delivery_point=point,
                order_index=idx,
            )
            for idx, point in enumerate(ordered_points)
        ]
        RouteSolutionDetail.objects.bulk_create(details)

    logger.debug(
        "save_solution | action=persist | route_id={route_id} | solution_id={solution_id} "
        "| details_count={count} | total_distance_km={dist}",
        route_id=route.id,
        solution_id=solution.id,
        count=len(details),
        dist=total_distance,
    )
    return solution

def process_route(route_id: int, task_id: str | None = None) -> None:
    """
    Punto de entrada principal. Ejecuta el flujo completo de procesamiento
    de una ruta y actualiza su estado en cada etapa.
    """
    start = time.perf_counter()

    try:
        route = Route.objects.select_related("input_file", "warehouse").get(id=route_id)
    except Route.DoesNotExist:
        logger.error(
            "process_route | action=route_not_found | result=failure | route_id={route_id} "
            "| task_id={task_id}",
            route_id=route_id,
            task_id=task_id,
        )
        return

    company_id = getattr(route.company, "id", None) if hasattr(route, "company") else None
    warehouse_id = getattr(route.warehouse, "id", None)

    logger.info(
        "process_route | action=start | route_id={route_id} | company_id={company_id} "
        "| warehouse_id={warehouse_id} | task_id={task_id}",
        route_id=route_id,
        company_id=company_id,
        warehouse_id=warehouse_id,
        task_id=task_id,
    )

    route.status = Status.PROCESSING
    route.save(update_fields=["status"])

    try:
        raw_rows = parse_input_file(route)
        logger.debug(
            "process_route | action=parse_done | route_id={route_id} | rows={rows}",
            route_id=route_id,
            rows=len(raw_rows),
        )

        resolved_rows = resolve_coordinates(raw_rows, route_id=route_id)
        logger.debug(
            "process_route | action=resolve_done | route_id={route_id} | resolved={resolved}",
            route_id=route_id,
            resolved=len(resolved_rows),
        )

        points = save_delivery_points(route, resolved_rows)
        logger.debug(
            "process_route | action=points_saved | route_id={route_id} | points={points}",
            route_id=route_id,
            points=len(points),
        )
 
        warehouse = route.warehouse
        ordered_points, total_distance = optimize_route(
            points,
            warehouse_lat=float(warehouse.latitude),
            warehouse_lng=float(warehouse.longitude),
            k_opt=route.k_opt,
        )
        logger.debug(
            "process_route | action=optimize_done | route_id={route_id} | total_distance_km={dist}",
            route_id=route_id,
            dist=total_distance,
        )

        save_solution(route, ordered_points, total_distance)
        elapsed_ms = int((time.perf_counter() - start) * 1000)

        logger.info(
            "process_route | action=end | result=success | route_id={route_id} "
            "| company_id={company_id} | delivery_count={delivery_count} "
            "| total_distance_km={dist} | execution_time_ms={elapsed_ms} | task_id={task_id}",
            route_id=route_id,
            company_id=company_id,
            delivery_count=len(points),
            dist=total_distance,
            elapsed_ms=elapsed_ms,
            task_id=task_id,
        )

        route.status = Status.COMPLETED

    except RouteProcessingError as exc:
        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.opt(exception=True).error(
            "process_route | action=end | result=controlled_error | route_id={route_id} "
            "| error_message={error} | execution_time_ms={elapsed_ms} | task_id={task_id}",
            route_id=route_id,
            error=str(exc),
            elapsed_ms=elapsed_ms,
            task_id=task_id,
        )
        route.status = Status.ERROR
        
    except Exception as exc:
        elapsed_ms = int((time.perf_counter() - start) * 1000)
        logger.opt(exception=True).critical(
            "process_route | action=end | result=unexpected_error | route_id={route_id} "
            "| error_message={error} | execution_time_ms={elapsed_ms} | task_id={task_id}",
            route_id=route_id,
            error=str(exc),
            elapsed_ms=elapsed_ms,
            task_id=task_id,
        )
        route.status = Status.ERROR

    finally:
        route.save(update_fields=["status"])